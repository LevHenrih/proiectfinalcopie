import openpyxl
from tkinter import *
from win32com import client
from openpyxl.chart import (PieChart, Reference)






root = Tk()
root.geometry("400x400")
root.resizable(0,0)
root.title("Script")
Label(root, text = "Test cases parser" , font  = "arial 15 bold").pack()
#root.mainloop()

tester_label = Label(root , text= "Tester Name" , font="arial 10 bold").pack()
tester_str = StringVar()
Entry(root, textvariable=tester_str).pack()
#excell:
path = r"C:\Users\camilo\Desktop\curstestare\proiect.manual\Test_Case_Format_[Levi].xlsx"
path_PDF = r"C:\Users\camilo\Desktop\curstestare\proiect.manual\Copy of Test_Case_Format_(Levi).xlsxTest_Case_Format_[Levi].pdf"
values = [0,0]
#total = 0
#failTestCases_counter = 0
#passTestCases_counter = 0
wb = openpyxl.load_workbook(path)


def compareValues():
    wb = openpyxl.load_workbook(path)
    first_sheet = wb["test case format"]
    #first_sheet = wb.get_sheet_by_name('test case format') #varianta care da eroare!
    total = 0
    #failTestCases_counter = 0
    #passTestCases_counter = 0
    for i in range(1, first_sheet.max_row):
        #print("i este,", i)
        if (first_sheet.cell(row=i, column= 7 ).value) == "FAILED":
            values[0]=values[0]+1
        elif (first_sheet.cell(row=i, column= 7 ).value) == "PASS":
            values[1] = values[1]+1
            totalvalue = values[0] + values[1]
    print("Total tests fail: ",values[0])
    print("Total tests pass: ",values[1])


    print("Total tests: ",totalvalue)

    #total = failTestCases_counter + passTestCases_counter
    #print('Total teste fail: ', failTestCases_counter)
    #print('Total teste pass: ', passTestCases_counter)
    #print('Total teste: ', total)

def createChart():
    wb = openpyxl.load_workbook(path)
    sheet = wb["Report"]
    pie = PieChart()

    labels = Reference(sheet,min_col= 1, min_row= 2,max_row=5)
    data = Reference(sheet,min_col= 2,min_row= 2,max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Test Cases"

    pie.width = 14
    pie.height = 7

    sheet.add_chart(pie,"A6")
    wb.save(path)


def generate_report():
    first_sheet = wb["test case format"]
    tester = first_sheet["E1"].value
    try:
        reportSheet = wb["Report"]
    except:

       wb.create_sheet("Report")
       reportSheet = wb["Report"]
    reportSheet["A1"] = "Tester ID"
    reportSheet["B1"] = tester
    reportSheet["A2"] = "Failed test cases"
    reportSheet["B2"] = values[0]
    reportSheet["A3"] = "Passed test cases"
    reportSheet["B3"] = values[1]
    reportSheet["A4"] = "Total number of test cases"
    reportSheet["B4"] = values[0] + values[1]



    wb.save(path)
    createChart()

    # generare pdf:
    excel = client.Dispatch("Excel.Application")
    # Read Excel File
    sheets = excel.Workbooks.Open(path)
    work_sheets = sheets.Worksheets[2]

    # Convert into PDF File
    work_sheets.ExportAsFixedFormat(0, path_PDF)


def buttonPressed():
    compareValues()
    generate_report()

Button(root, text="Generate report", command=buttonPressed).pack(pady=10)

root.mainloop()

#compareValues()


#generate_report()
